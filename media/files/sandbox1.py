import csv
import os
import glob
from openpyxl import load_workbook

# Get the path to the media/files directory
files_directory = os.path.join(os.getcwd(), 'media', 'files')

# Find any CSV file in the media/files directory
csv_files = glob.glob(os.path.join(files_directory, '*.csv'))

# Find any Excel file in the media/files directory
xlsx_files = glob.glob(os.path.join(files_directory, '*.xlsx'))

# If there are any CSV files, open the most recent one
if csv_files and xlsx_files:
    most_recent_csv_file = max(csv_files, key=os.path.getctime)  # Get the most recent CSV file based on creation time
    most_recent_xlsx_file = max(xlsx_files, key=os.path.getctime)  # Get the most recent Excel file based on creation time

    with open(most_recent_csv_file, 'r') as csv_file:
        reader = csv.reader(csv_file)
        data = list(reader)

    # Load the Excel workbook and select the active sheet
    workbook = load_workbook(filename=most_recent_xlsx_file)
    sheet = workbook.active

    # Write the data to the Excel file starting from cell A2
    for i, row in enumerate(data, start=2):
        for j, cell in enumerate(row, start=1):
            sheet.cell(row=i, column=j, value=cell)

    # Save the changes to the Excel file
    workbook.save(most_recent_xlsx_file)

    print("Excel file updated successfully:", most_recent_xlsx_file)
else:
    print("No CSV or Excel files found in the media/files directory.")

    


